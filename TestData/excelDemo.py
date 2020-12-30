import openpyxl

book = openpyxl.load_workbook("C:\\Users\\brahm\\Desktop\\Python\\Udemy\\Rahul Shetty\\pythonSelFramework\\TestData\\PythonDemo.xlsx")
sheet = book.active

#cell = sheet.cell(row=1, column=2)
#print(cell.value)

#testcase1_first_name = sheet.cell(row=2, c olumn=2).value = "Tito"
#print(testcase1_first_name)

#print(sheet.max_row)

#print(sheet['A5'].value)

#for i in range(1, sheet.max_row+1):
#    for j in range(1, sheet.max_column+1):
#        print(sheet.cell(row=i, column=j).value)
		
		
#for i in range(1, sheet.max_row+1):
#    if sheet.cell(row=i, column=1).value == "Testcase2":
#        for j in range(1, sheet.max_column+1):
#            print(sheet.cell(row=i, column=j).value)
			

#for i in range(1, sheet.max_row+1):
#    if sheet.cell(row=i, column=1).value == "Testcase2":
#        for j in range(2, sheet.max_column+1):
#            print(sheet.cell(row=i, column=j).value)

dict = {}
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)